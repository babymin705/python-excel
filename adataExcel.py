#!/usr/bin/python3
import mysql.connector
# from os import walk
import os
from pathlib import Path
import datetime
from dateutil.relativedelta import relativedelta
import openpyxl
cnx = mysql.connector.connect(user='root', password='ADATA@STVH0tel',host='127.0.0.1',database='adata')


# Read the files in the folder
f = []
mypath = '/var/www/adata/adata_stats/'
for(dirpath, dirnames, filenames) in os.walk(mypath):
    # p = Path(filenames)
    # f.extend(filenames)
    for name in filenames:
        f.append((os.path.join(dirpath, name)))

for filename in f:
    p = Path('/var/www/adata/adata_stats/',filename)
    # Create the unix timestamp based on file name
    name = str(p.stem).split('-')
    first_day = datetime.datetime(int(name[0]), int(name[1]), 1) #first day
    last_day = first_day + relativedelta(hour=23, minute=59, second=00, day=31) #last day
    unix_first_day_in_month = int(round(first_day.timestamp()))
    unix_last_day_in_month = int(round(last_day.timestamp()))
    # Read all the data in the excel file
    wb_obj = openpyxl.load_workbook(p)
    # sheet = wb_obj.active

    # col_names = []
    # for column in sheet.iter_cols(min_row=1, max_col=sheet.max_column):
    #     col_names.append(column[0].value)
    # print(col_names)
    # data = {'hotel':[], 'occ':[], 'arr':[]}
    keywords = ["aor","arr","rooms_available"]
    print("*** start file %s ****"%(filename))
    for sheet in wb_obj.worksheets:
        print("*** start sheet %s ****"%(sheet))
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i != 0:
                if(row[0]):
                    hotel_name = row[0]
                    hotel_aor = row[1].strip() #average_occupancy_rate
                    hotel_arr = row[2].strip() #average_room_rate
                    # skip if hotel_arr and hotel_aor is 0
                    if(hotel_aor != "0.00" and hotel_arr != "0.00"):
                        # Search the hotel name in hotels table
                        search_hotel_query = ("SELECT id, rooms FROM hotels WHERE name = %s")
                        cursor = cnx.cursor()
                        cursor.execute(search_hotel_query, (hotel_name,))
                        hotel = cursor.fetchone()
                        # get the total rooms for available rooms purpose
                        if(hotel):
                            # check in stats table
                            search_stat_query = ('SELECT * FROM stats where hotel_id=%s and start_date = %s')
                            stat_data = (hotel[0], unix_first_day_in_month)
                            cursor.execute(search_stat_query, stat_data)
                            stat = cursor.fetchone()
                            if(not stat):
                                hotel_room_available = hotel[1]
                                hotel_id = hotel[0]
                                hotel_data = [float(hotel_aor), float(hotel_arr), float(hotel_room_available)]
                                # insert the data inside the stats table
                                stats_data = (hotel_id,unix_first_day_in_month,unix_last_day_in_month)
                                add_stats_query = ("INSERT INTO stats (hotel_id, start_date, end_date, send, updated) "
                                                    "VALUES (%s,%s,%s,1,1)")
                                cursor.execute(add_stats_query, stats_data)
                                stats_id = cursor.lastrowid
                                # add to stat_details
                                for i, keyword in enumerate(keywords):
                                    add_stats_detail_query = ("INSERT INTO stat_details (stats_id, keyword, value) "
                                                                "VALUES (%s, %s, %s)")
                                    stats_detail = (stats_id, keyword, hotel_data[i])
                                    cursor.execute(add_stats_detail_query, stats_detail)
                            
                        else:
                            print("*** missing hotel %s"%hotel_name)

cursor.close()
cnx.commit()
cnx.close()