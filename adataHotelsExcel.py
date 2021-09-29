#!/usr/bin/python3
import mysql.connector
from os import walk
from pathlib import Path
import datetime
from dateutil.relativedelta import relativedelta
import openpyxl
cnx = mysql.connector.connect(user='root', password='ADATA@STVH0tel',host='127.0.0.1',database='adata')

cursor = cnx.cursor(buffered=True)
# Read the files in the folder
f = []
mypath = '/var/www/adata/adata_hotels'
for(dirpath, dirnames, filenames) in walk(mypath):
    # p = Path(filenames)
    f.extend(filenames)

for filename in filenames:
    p = Path('/var/www/adata/adata_hotels',filename)
    # Read all the data in the excel file
    wb_obj = openpyxl.load_workbook(p)
    # sheet = wb_obj.active
    for sheet in wb_obj.worksheets:
        print("*** start sheet %s ****"%(sheet))
        for i, row in enumerate(sheet.iter_rows(values_only=True)):
            if i != 0:
                if(row[5]):
                    hotel_name = row[0].strip()
                    hotel_state = row[1].strip()
                    hotel_city = row[2].strip()
                    hotel_star = row[3]
                    hotel_roomno = row[4]
                    hotel_ota_hotel = row[5].strip()
                    # Search the hotel name in hotels table
                    search_hotel_query = ('SELECT * from hotels where name =%s')
                    cursor.execute(search_hotel_query, (hotel_name,))
                    hotel = cursor.fetchone()
                    if(not hotel):
                        # Search ota hotel name in ota_hotels table
                        search_ota_hotel_query = ("SELECT * FROM ota_hotels WHERE name like %s")
                        
                        cursor.execute(search_ota_hotel_query, ("%"+hotel_ota_hotel+"%",))
                        item = cursor.fetchone()
                        if(item):
                            # insert_cursor = cnx.cursor(buffered=True)
                                # insert the data inside the hotels table
                            add_hotel_query = ("INSERT INTO hotels (name, ota_hotel_id, addr, city, state, country_code_id, phone, email,whatsapp_no, rooms, stars, active) "
                                                "VALUES (%s, %s, 'pilot_hotel', %s, %s, 1, '0109999999', 'pilot_hotel@hotel.com','01099999999',%s,%s,1)")
                            data_hotel = (hotel_name, item[0], hotel_city,hotel_state,hotel_roomno,hotel_star)
                            cursor.execute(add_hotel_query, data_hotel)
                        else:
                            print("*** cannot find ota hotel %s"%(hotel_ota_hotel))
                            add_hotel_query = ("INSERT INTO hotels (name,  addr, city, state, country_code_id, phone, email,whatsapp_no, rooms, stars, active) "
                                                "VALUES (%s, 'pilot_hotel', %s, %s, 1, '0109999999', 'pilot_hotel@hotel.com','01099999999',%s,%s,1)")
                            data_hotel = (hotel_name, hotel_city,hotel_state,hotel_roomno,hotel_star)
                            cursor.execute(add_hotel_query, data_hotel)
                    else:
                        print("*** hotel exist %s"%(hotel_name))
                        # update the hotel
                        search_ota_hotel_query = ("SELECT * FROM ota_hotels WHERE name like %s")
                        cursor.execute(search_ota_hotel_query, ("%"+hotel_ota_hotel+"%",))
                        item = cursor.fetchone()
                        if(item):
                            update_hotel_query = ("UPDATE hotels set ota_hotel_id = %s, city = %s, state = %s, rooms = %s, stars = %s where id = %s")
                            data_hotel = (item[0], hotel_city,hotel_state,hotel_roomno,hotel_star, hotel[0])
                        else:
                            update_hotel_query = ("UPDATE hotels set city = %s, state = %s, rooms = %s, stars = %s where id = %s")
                            data_hotel = (hotel_city,hotel_state,hotel_roomno,hotel_star, hotel[0])
                        cursor.execute(update_hotel_query, data_hotel)
                        cursor.reset()

        print("**** end sheet *****")


    # col_names = []
    # for column in sheet.iter_cols(min_row=1, max_col=sheet.max_column):
    #     col_names.append(column[0].value)
    # print(col_names)
    # data = {'hotel':[], 'occ':[], 'arr':[]}
    # for i, row in enumerate(sheet.iter_rows(values_only=True)):
    #     if i != 0:
    #         if(row[3]):
    #             hotel_name = row[3]
    #             hotel_occ = row[1].strip()
    #             hotel_arr = row[2].strip()
    #             print(hotel_name)
    #             # Search the hotel name in hotels table
    #             search_ota_hotel_query = ("SELECT * FROM hotels WHERE name = %s")
    #             cursor = cnx.cursor()
    #             cursor.execute(search_ota_hotel_query, (hotel_name,))
    #             for(item) in cursor:
    #                 print(item)
    #                 # insert the data inside the hotels table
    #                 stats_data = (item[0],unix_first_day_in_month,unix_last_day_in_month,1,1)
    #                 # add_stats_query = ("INSERT INTO stats () "
    #                 #                     "VALUES ()")




# insert the monthly report inside the hotels table
cursor.close()
cnx.commit()
cnx.close()